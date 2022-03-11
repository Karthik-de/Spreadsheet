import openpyxl

# Is use to create a reference of the Excel to wb
wb1 = openpyxl.load_workbook('test_data.xlsx')
wb2 = openpyxl.load_workbook('test_data_3.xlsx')

# Refrence the workbook to the worksheets
sh1 = wb1.active
sh2 = wb2["Sheet1"]  # use same sheet name, different workbook

for row in sh1.iter_rows():
    if row[6].value == 16:
        sh2.append((cell.value for cell in row))

wb1.save("test_data.xlsx")
wb2.save("test_data_3.xlsx")
